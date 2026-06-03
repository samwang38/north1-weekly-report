#!/usr/bin/env python3
"""北一區週報產生器 — 本機 Web 工具"""
import json, os, sys, time, threading, traceback, urllib.parse, uuid
from calendar import monthrange
from datetime import date, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path

ROOT         = Path(__file__).resolve().parent
STATIC_ROOT  = ROOT / 'static'
TEMPLATE     = ROOT / 'template' / '北一區週報_優化.xlsx'
SA_FILE      = ROOT / 'data' / 'SAcare對應價目表.xlsx'
TRAFFIC_FILE = ROOT / 'data' / 'traffic_cache.json'

sys.path.insert(0, str(ROOT))
import multistore_engine as eng
from openpyxl import load_workbook

JOBS = {}
_LOCK = threading.Lock()


# ─── 人流快取（插件背景推送）─────────────────────────────────
def _load_traffic() -> dict:
    """讀取人流快取 → {storeCode: {'YYYY-MM-DD': visitors}}"""
    try:
        with open(TRAFFIC_FILE, encoding='utf-8') as f:
            return json.load(f).get('stores', {})
    except Exception:
        return {}


def _save_traffic(stores: dict):
    TRAFFIC_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(TRAFFIC_FILE, 'w', encoding='utf-8') as f:
        json.dump({'updated': time.strftime('%Y-%m-%d %H:%M:%S'),
                   'stores': stores}, f, ensure_ascii=False)


def _traffic_sum(stores: dict, code: str, start: date, end: date) -> int:
    """加總某店在 [start, end] 區間的每日人流"""
    days = stores.get(code, {})
    s, e = start.isoformat(), end.isoformat()
    return int(sum(v for d, v in days.items() if s <= d <= e))


# ─── 預設本週結束日（最近一個週六）─────────────────────────────
def _last_saturday() -> date:
    today = date.today()
    days = (today.weekday() + 2) % 7   # Sat→0, Sun→1, Mon→2, …
    return today - timedelta(days=days if days else 7)


# ─── 核心填充邏輯 ─────────────────────────────────────────────
def _fill_workbook(wk_end: date, log, use_full_month: bool = False) -> bytes:
    from openpyxl import load_workbook as _lw
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font as _Font

    # 保固合計搭售率目標（合計搭售率欄 → 目標%）：未達標轉紅字
    WARRANTY_TARGETS = {7: 0.60, 13: 0.40, 19: 0.50, 25: 0.19, 32: 0.35}
    _RED_FONT = _Font(color='FFFF0000')

    WK_START = wk_end - timedelta(days=6)
    PW_END   = WK_START - timedelta(days=1)
    PW_START = PW_END - timedelta(days=6)

    # MTD（跨月時用 WK_START 那個月；use_full_month 時延伸到月底）
    if WK_START.month != wk_end.month:
        MTD_START = date(WK_START.year, WK_START.month, 1)
        MTD_END   = date(WK_START.year, WK_START.month,
                         monthrange(WK_START.year, WK_START.month)[1])
    else:
        MTD_START = date(wk_end.year, wk_end.month, 1)
        if use_full_month:
            MTD_END = date(wk_end.year, wk_end.month,
                           monthrange(wk_end.year, wk_end.month)[1])
        else:
            MTD_END = wk_end

    _pm_y  = MTD_START.year if MTD_START.month > 1 else MTD_START.year - 1
    _pm_m  = MTD_START.month - 1 if MTD_START.month > 1 else 12
    PM_SAME_START = date(_pm_y, _pm_m, 1)
    PM_SAME_END   = date(_pm_y, _pm_m, min(MTD_END.day, monthrange(_pm_y, _pm_m)[1]))
    PM_START = PM_SAME_START
    PM_END   = date(_pm_y, _pm_m, monthrange(_pm_y, _pm_m)[1])

    LYMO_START = date(MTD_START.year - 1, MTD_START.month, 1)
    LYMO_END   = date(LYMO_START.year, LYMO_START.month,
                      min(MTD_END.day, monthrange(LYMO_START.year, LYMO_START.month)[1]))

    LYW_START = WK_START - timedelta(weeks=52)
    LYW_END   = wk_end   - timedelta(weeks=52)
    YTD_S_CY  = date(MTD_END.year, 1, 1);  YTD_E_CY = MTD_END
    YTD_S_LY  = date(LYMO_END.year, 1, 1); YTD_E_LY = LYMO_END

    STORE_CODES = list(eng.STORES.keys())

    log(f'本週: {WK_START} ~ {wk_end}  |  上週: {PW_START} ~ {PW_END}')
    log(f'MTD: {MTD_START}~{MTD_END}  PM_SAME: {PM_SAME_START}~{PM_SAME_END}')
    log(f'LYMO: {LYMO_START}~{LYMO_END}')

    log('載入今年銷售資料（EPB）…')
    df_cy = eng.load_from_epb(YTD_S_CY, YTD_E_CY, store_codes=STORE_CODES)
    log(f'  今年資料：{len(df_cy):,} 筆')

    ly_end = max(LYW_END, LYMO_END, YTD_E_LY)
    log('載入去年銷售資料（EPB）…')
    df_ly = eng.load_from_epb(YTD_S_LY, ly_end, store_codes=STORE_CODES)
    log(f'  去年資料：{len(df_ly):,} 筆')

    sa_prices = eng.load_sacare_prices(str(SA_FILE))
    sa_codes  = set(sa_prices.keys())

    log('計算各期指標…')

    def rate(num, den):
        return round(num / den, 4) if den else 0.0

    def pct(new, old):
        return round((new - old) / abs(old), 4) if old else 0.0

    PERIODS = {
        'wk':      (df_cy, WK_START,      wk_end),
        'pw':      (df_cy, PW_START,      PW_END),
        'mtd':     (df_cy, MTD_START,     MTD_END),
        'pm_same': (df_cy, PM_SAME_START, PM_SAME_END),
        'pm':      (df_cy, PM_START,      PM_END),
        'lyw':     (df_ly, LYW_START,     LYW_END),
        'lymo':    (df_ly, LYMO_START,    LYMO_END),
        'ytd_cy':  (df_cy, YTD_S_CY,      YTD_E_CY),
        'ytd_ly':  (df_ly, YTD_S_LY,      YTD_E_LY),
    }
    M = {}; A = {}
    rows_stores = STORE_CODES + ['ALL']
    for period, (df, s, e) in PERIODS.items():
        M[period] = {}; A[period] = {}
        for code in rows_stores:
            sc = None if code == 'ALL' else code
            M[period][code] = eng.calc_store_metrics(df, s, e, sc, sa_prices)
            A[period][code] = eng.calc_accessory_by_c4(df, s, e, sc, sa_codes)

    MISC_WK = {}; MISC_PW = {}; MISC_MTD = {}
    for code in rows_stores:
        sc = None if code == 'ALL' else code
        MISC_WK[code]  = eng.calc_misc_metrics(df_cy, WK_START, wk_end,    sc, sa_prices)
        MISC_PW[code]  = eng.calc_misc_metrics(df_cy, PW_START, PW_END,    sc, sa_prices)
        MISC_MTD[code] = eng.calc_misc_metrics(df_cy, MTD_START, MTD_END,  sc, sa_prices)

    # 人流：依期間加總每店每日人流（羅東無計數器，維持公式不填）
    _traffic = _load_traffic()
    TRAFFIC = {}
    for period, (df, s, e) in PERIODS.items():
        TRAFFIC[period] = {c: _traffic_sum(_traffic, c, s, e) for c in STORE_CODES}
    if _traffic:
        log(f'人流快取：{len(_traffic)} 店有資料')
    else:
        log('人流快取：無（人流欄留空，可由插件背景推送）')

    log('指標計算完成，載入範本…')
    wb = _lw(str(TEMPLATE))

    # ── 格式輔助 ──────────────────────────────────────────────────
    FMT_INT = '#,##0'
    FMT_PCT = '0.00%'

    def _set(ws, r, c, val, fmt=FMT_INT):
        cell = ws.cell(r, c)
        cell.value = val
        cell.number_format = fmt

    def _fml_diff(ws, r, c_new, c_old, r_ref=None):
        """差值公式：=NEW-OLD，整數格式"""
        rr = r_ref or r
        cell = ws.cell(r, c_new + 1 if False else c_new)
        # 直接用 offset 計算 formula col
        new_ref = f'{_gcl(c_new)}{rr}'
        old_ref = f'{_gcl(c_old)}{rr}'
        ws.cell(r, c_new).value = f'={new_ref}-{old_ref}'
        ws.cell(r, c_new).number_format = FMT_INT

    def _fml_pct(ws, r, c_out, c_new, c_old):
        """百分比公式：=(NEW-OLD)/ABS(OLD)，百分比格式"""
        new_ref = f'{_gcl(c_new)}{r}'
        old_ref = f'{_gcl(c_old)}{r}'
        ws.cell(r, c_out).value = f'=IF({old_ref}=0,0,({new_ref}-{old_ref})/ABS({old_ref}))'
        ws.cell(r, c_out).number_format = FMT_PCT

    def _fml_rate(ws, r, c_out, c_num, c_den):
        """比率公式：=NUM/DEN，百分比格式"""
        num_ref = f'{_gcl(c_num)}{r}'
        den_ref = f'{_gcl(c_den)}{r}'
        ws.cell(r, c_out).value = f'=IF({den_ref}=0,0,{num_ref}/{den_ref})'
        ws.cell(r, c_out).number_format = FMT_PCT

    def _fml_sum(ws, r_total, c, r_start, r_end):
        """SUM 公式：加總 r_start~r_end 列，整數格式"""
        col_l = _gcl(c)
        ws.cell(r_total, c).value = f'=SUM({col_l}{r_start}:{col_l}{r_end})'
        ws.cell(r_total, c).number_format = FMT_INT

    # ── helpers ──────────────────────────────────────────────────
    def fill_acc_section(ws, row_start, col_offset, code):
        n = len(eng.C4_ACCESSORY)
        for i, (c4, _) in enumerate(eng.C4_ACCESSORY):
            r = row_start + i
            o = col_offset  # shorthand
            _set(ws, r, o+3,  A['pw'][code].get(c4, 0))
            _set(ws, r, o+4,  A['wk'][code].get(c4, 0))
            _set(ws, r, o+7,  A['mtd'][code].get(c4, 0))
            _set(ws, r, o+8,  A['pm'][code].get(c4, 0))
            _set(ws, r, o+11, A['lyw'][code].get(c4, 0))
            _set(ws, r, o+14, A['ytd_ly'][code].get(c4, 0))
            _set(ws, r, o+15, A['ytd_cy'][code].get(c4, 0))
            # 差值與百分比→公式
            ws.cell(r, o+5).value = f'={_gcl(o+4)}{r}-{_gcl(o+3)}{r}'
            ws.cell(r, o+5).number_format = FMT_INT
            _fml_pct(ws, r, o+6,  o+4, o+3)
            ws.cell(r, o+9).value = f'={_gcl(o+7)}{r}-{_gcl(o+8)}{r}'
            ws.cell(r, o+9).number_format = FMT_INT
            _fml_pct(ws, r, o+10, o+7, o+8)
            ws.cell(r, o+12).value = f'={_gcl(o+4)}{r}-{_gcl(o+11)}{r}'
            ws.cell(r, o+12).number_format = FMT_INT
            _fml_pct(ws, r, o+13, o+4, o+11)
            ws.cell(r, o+16).value = f'={_gcl(o+15)}{r}-{_gcl(o+14)}{r}'
            ws.cell(r, o+16).number_format = FMT_INT
            _fml_pct(ws, r, o+17, o+15, o+14)

        total_r = row_start + n
        r_s = row_start; r_e = row_start + n - 1
        for off in [3, 4, 7, 8, 11, 14, 15]:
            _fml_sum(ws, total_r, col_offset+off, r_s, r_e)
        o = col_offset
        ws.cell(total_r, o+5).value = f'={_gcl(o+4)}{total_r}-{_gcl(o+3)}{total_r}'
        ws.cell(total_r, o+5).number_format = FMT_INT
        _fml_pct(ws, total_r, o+6,  o+4, o+3)
        ws.cell(total_r, o+9).value = f'={_gcl(o+7)}{total_r}-{_gcl(o+8)}{total_r}'
        ws.cell(total_r, o+9).number_format = FMT_INT
        _fml_pct(ws, total_r, o+10, o+7, o+8)
        ws.cell(total_r, o+12).value = f'={_gcl(o+4)}{total_r}-{_gcl(o+11)}{total_r}'
        ws.cell(total_r, o+12).number_format = FMT_INT
        _fml_pct(ws, total_r, o+13, o+4, o+11)
        ws.cell(total_r, o+16).value = f'={_gcl(o+15)}{total_r}-{_gcl(o+14)}{total_r}'
        ws.cell(total_r, o+16).number_format = FMT_INT
        _fml_pct(ws, total_r, o+17, o+15, o+14)

    def fill_acpp(ws, row_start, period):
        for ri, code in enumerate(rows_stores):
            r = row_start + ri
            m = M[period][code]
            ws.cell(r, 1).value = eng.STORES.get(code, 'Total')
            _set(ws, r,  2, m['cpu_units'])
            _set(ws, r,  3, m['acpp_mac'])
            _fml_rate(ws, r,  4, 3, 2)
            _set(ws, r,  5, m['sa_cpu'])
            _fml_rate(ws, r,  6, 5, 2)
            # col7: (ACPP+SA)/cpu = (col3+col5)/col2
            ws.cell(r, 7).value = f'=IF(B{r}=0,0,(C{r}+E{r})/B{r})'
            ws.cell(r, 7).number_format = FMT_PCT
            _set(ws, r,  8, m['watch_units'])
            _set(ws, r,  9, m['acpp_watch'])
            _fml_rate(ws, r, 10, 9, 8)
            _set(ws, r, 11, m['sa_watch'])
            _fml_rate(ws, r, 12, 11, 8)
            ws.cell(r,13).value = f'=IF(H{r}=0,0,(I{r}+K{r})/H{r})'
            ws.cell(r,13).number_format = FMT_PCT
            _set(ws, r, 14, m['ipad_units'])
            _set(ws, r, 15, m['acpp_ipad'])
            _fml_rate(ws, r, 16, 15, 14)
            _set(ws, r, 17, m['sa_ipad'])
            _fml_rate(ws, r, 18, 17, 14)
            ws.cell(r,19).value = f'=IF(N{r}=0,0,(O{r}+Q{r})/N{r})'
            ws.cell(r,19).number_format = FMT_PCT
            _set(ws, r, 20, m['iphone_units'])
            _set(ws, r, 21, m['acpp_iphone'])
            _fml_rate(ws, r, 22, 21, 20)
            _set(ws, r, 23, m['sa_iphone'])
            _fml_rate(ws, r, 24, 23, 20)
            ws.cell(r,25).value = f'=IF(T{r}=0,0,(U{r}+W{r})/T{r})'
            ws.cell(r,25).number_format = FMT_PCT
            # col26: iOS搭售率 = (iPhone+iPad ACPP+SA) / (iPhone+iPad台數)
            ws.cell(r,26).value = f'=IF(T{r}+N{r}=0,0,(U{r}+W{r}+O{r}+Q{r})/(T{r}+N{r}))'
            ws.cell(r,26).number_format = FMT_PCT
            _set(ws, r, 27, m['airpods_units'])
            _set(ws, r, 28, m['acpp_airpods'])
            _fml_rate(ws, r, 29, 28, 27)
            _set(ws, r, 30, m['sa_airpods'])
            _fml_rate(ws, r, 31, 30, 27)
            ws.cell(r,32).value = f'=IF(AA{r}=0,0,(AB{r}+AD{r})/AA{r})'
            ws.cell(r,32).number_format = FMT_PCT
        # 合計搭售率未達目標 → 紅字（條件式格式，依儲存格值動態套用）
        r_end = row_start + len(rows_stores) - 1
        for col, tgt in WARRANTY_TARGETS.items():
            col_l = _gcl(col)
            ws.conditional_formatting.add(
                f'{col_l}{row_start}:{col_l}{r_end}',
                CellIsRule(operator='lessThan', formula=[str(tgt)], font=_RED_FONT))

    def fill_units(ws, row_start, pa, pb):
        for ri, code in enumerate(rows_stores):
            r = row_start + ri
            ma = M[pa][code]; mb = M[pb][code]
            ws.cell(r, 1).value = eng.STORES.get(code, 'Total')
            _set(ws, r,  2, ma['cpu_units'])
            _set(ws, r,  3, mb['cpu_units'])
            _fml_pct(ws, r, 4, 3, 2)
            _set(ws, r,  5, ma['ipad_units'])
            _set(ws, r,  6, mb['ipad_units'])
            _fml_pct(ws, r, 7, 6, 5)
            _set(ws, r,  8, ma['iphone_units'])
            _set(ws, r,  9, mb['iphone_units'])
            _fml_pct(ws, r, 10, 9, 8)
            _set(ws, r, 11, ma['watch_units'])
            _set(ws, r, 12, mb['watch_units'])
            _fml_pct(ws, r, 13, 12, 11)
            _set(ws, r, 14, ma['airpods_units'])
            _set(ws, r, 15, mb['airpods_units'])
            _fml_pct(ws, r, 16, 15, 14)
            _set(ws, r, 18, ma['txn_count'])     # 成交筆數 上期 (R)
            _set(ws, r, 21, mb['txn_count'])     # 成交筆數 本期 (U)
            # ── 人流 (S=上期, V=本期) ──
            store_last = row_start + len(rows_stores) - 2   # 最後一個門市列（羅東）
            if code == '057':
                # 羅東無計數器：人流 = 成交筆數 × 0.85 / 0.3（去小數）
                ws.cell(r, 19).value = f'=ROUND(R{r}*0.85/0.3,0)'
                ws.cell(r, 19).number_format = FMT_INT
                ws.cell(r, 22).value = f'=ROUND(U{r}*0.85/0.3,0)'
                ws.cell(r, 22).number_format = FMT_INT
            elif code == 'ALL':
                # Total：人流加總
                _fml_sum(ws, r, 19, row_start, store_last)
                _fml_sum(ws, r, 22, row_start, store_last)
            else:
                # 5 店：有人流快取就填，沒有就留空白（手填）
                ta = TRAFFIC[pa].get(code, 0)
                tb = TRAFFIC[pb].get(code, 0)
                if ta: _set(ws, r, 19, ta)
                if tb: _set(ws, r, 22, tb)
            # 提袋率 = 成交筆數 / 人流
            _fml_rate(ws, r, 20, 18, 19)         # T = R/S
            _fml_rate(ws, r, 23, 21, 22)         # W = U/V
            # 差異人流 = 本期人流 - 上期人流
            ws.cell(r, 24).value = f'=V{r}-S{r}'
            ws.cell(r, 24).number_format = FMT_INT
            # 較上期比 = (本期-上期)/上期
            _fml_pct(ws, r, 25, 22, 19)          # Y = (V-S)/S
            # ── 平均單價 = 業績/筆數 ──
            avg_a = rate(ma['total_excl_sa'], ma['txn_count'])
            avg_b = rate(mb['total_excl_sa'], mb['txn_count'])
            _set(ws, r, 26, round(avg_a))
            _set(ws, r, 27, round(avg_b))
            ws.cell(r, 28).value = f'={_gcl(27)}{r}-{_gcl(26)}{r}'
            ws.cell(r, 28).number_format = FMT_INT
            _fml_pct(ws, r, 29, 27, 26)

    def fill_biz(ws, row_start, pa, pb):
        for ri, code in enumerate(rows_stores):
            r = row_start + ri
            a = M[pa][code]; b = M[pb][code]
            ws.cell(r, 1).value = eng.STORES.get(code, 'Total')
            _set(ws, r,  2, a['total_excl_sa'])
            _set(ws, r,  3, b['total_excl_sa'])
            _fml_pct(ws, r,  4, 3, 2)
            _set(ws, r,  5, a['tpp_excl_sa'])
            _set(ws, r,  6, b['tpp_excl_sa'])
            _fml_pct(ws, r,  7, 6, 5)
            _set(ws, r,  8, a['sa_rev'])
            _set(ws, r,  9, b['sa_rev'])
            _fml_pct(ws, r, 10, 9, 8)
            _set(ws, r, 11, a['acpp_plus'])
            _set(ws, r, 12, b['acpp_plus'])
            _fml_pct(ws, r, 13, 12, 11)
            _set(ws, r, 14, a['total_rev'])
            _set(ws, r, 15, b['total_rev'])
            _fml_pct(ws, r, 16, 15, 14)
            _set(ws, r, 17, a['tpp_rev'])
            _set(ws, r, 18, b['tpp_rev'])
            _fml_pct(ws, r, 19, 18, 17)
            ws.cell(r, 20).value = f'={_gcl(3)}{r}-{_gcl(2)}{r}'
            ws.cell(r, 20).number_format = FMT_INT
            ws.cell(r, 21).value = f'={_gcl(6)}{r}-{_gcl(5)}{r}'
            ws.cell(r, 21).number_format = FMT_INT
            ws.cell(r, 22).value = f'={_gcl(9)}{r}-{_gcl(8)}{r}'
            ws.cell(r, 22).number_format = FMT_INT
            ws.cell(r, 23).value = f'={_gcl(12)}{r}-{_gcl(11)}{r}'
            ws.cell(r, 23).number_format = FMT_INT
            _fml_rate(ws, r, 24, 5, 2)
            _fml_rate(ws, r, 25, 6, 3)
            _fml_rate(ws, r, 26, 8, 2)
            _fml_rate(ws, r, 27, 9, 3)
            _set(ws, r, 28, a['coupon_rev'])
            _set(ws, r, 29, b['coupon_rev'])
            _fml_pct(ws, r, 30, 29, 28)

    def fill_biz_mo(ws, row_start, pa, pb):
        """月累積版（多兩個手動欄 col2-3）"""
        for ri, code in enumerate(rows_stores):
            r = row_start + ri
            a = M[pa][code]; b = M[pb][code]
            ws.cell(r, 1).value = eng.STORES.get(code, 'Total')
            _set(ws, r,  4, a['total_excl_sa'])
            _set(ws, r,  5, b['total_excl_sa'])
            _fml_pct(ws, r,  6, 5, 4)
            _set(ws, r,  7, a['tpp_excl_sa'])
            _set(ws, r,  8, b['tpp_excl_sa'])
            _fml_pct(ws, r,  9, 8, 7)
            _set(ws, r, 10, a['sa_rev'])
            _set(ws, r, 11, b['sa_rev'])
            _fml_pct(ws, r, 12, 11, 10)
            _set(ws, r, 13, a['acpp_plus'])
            _set(ws, r, 14, b['acpp_plus'])
            _fml_pct(ws, r, 15, 14, 13)
            _set(ws, r, 16, a['total_rev'])
            _set(ws, r, 17, b['total_rev'])
            _fml_pct(ws, r, 18, 17, 16)
            _set(ws, r, 19, a['tpp_rev'])
            _set(ws, r, 20, b['tpp_rev'])
            _fml_pct(ws, r, 21, 20, 19)
            ws.cell(r, 22).value = f'={_gcl(5)}{r}-{_gcl(4)}{r}'
            ws.cell(r, 22).number_format = FMT_INT
            ws.cell(r, 23).value = f'={_gcl(8)}{r}-{_gcl(7)}{r}'
            ws.cell(r, 23).number_format = FMT_INT
            # SA Care差異 = 本月 - 上月 (K-J)
            ws.cell(r, 24).value = f'={_gcl(11)}{r}-{_gcl(10)}{r}'
            ws.cell(r, 24).number_format = FMT_INT
            # AC差異 = 本月 - 上月 (N-M)
            ws.cell(r, 25).value = f'={_gcl(14)}{r}-{_gcl(13)}{r}'
            ws.cell(r, 25).number_format = FMT_INT
            # 達成率 = 本月總業績 / 本月目標 (Q/B)
            _fml_rate(ws, r, 3, 17, 2)
            _fml_rate(ws, r, 26, 7, 4)
            _fml_rate(ws, r, 27, 8, 5)
            _fml_rate(ws, r, 28, 10, 4)
            _fml_rate(ws, r, 29, 11, 5)
            _set(ws, r, 30, a['coupon_rev'])
            _set(ws, r, 31, b['coupon_rev'])
            _fml_pct(ws, r, 32, 31, 30)

    # ── 配件 sheets ──────────────────────────────────────────────
    log('填入配件 sheets…')
    ws_sum = wb['配件-北一區 (匯總)']
    fill_acc_section(ws_sum, 4, 0, 'ALL')
    fill_acc_section(ws_sum, 23, 0,  '004'); fill_acc_section(ws_sum, 23, 17, '054')
    fill_acc_section(ws_sum, 42, 0,  '024'); fill_acc_section(ws_sum, 42, 17, '046')
    fill_acc_section(ws_sum, 61, 0,  '005'); fill_acc_section(ws_sum, 61, 17, '057')
    fill_acc_section(wb['配件-北一區'], 4, 0, 'ALL')
    for code, sname in [('004','配件 - 士林門市'), ('005','配件 - 微風門市'),
                        ('024','配件 - 美麗華門市'), ('046','配件 - 阿波羅門市'),
                        ('054','配件 - 大葉高島屋門市'), ('057','配件 - 羅東門市')]:
        fill_acc_section(wb[sname], 4, 0, code)

    # ── BY店 本週比較 ─────────────────────────────────────────────
    log('填入 BY店 本週比較…')
    ws_wk = wb['BY店 本週比較']
    fill_biz(ws_wk, 4, 'pw', 'wk')
    fill_units(ws_wk, 14, 'pw', 'wk')
    fill_acpp(ws_wk, 23, 'wk');  fill_acpp(ws_wk, 32, 'pw')

    # ── BY店 月累積 ───────────────────────────────────────────────
    log('填入 BY店 月累積…')
    ws_mo = wb['BY店 月累積']
    fill_biz_mo(ws_mo, 4, 'pm_same', 'mtd')
    fill_units(ws_mo, 14, 'pm_same', 'mtd')
    fill_acpp(ws_mo, 23, 'mtd');  fill_acpp(ws_mo, 32, 'pm_same')

    # ── BY店 去年同期 ─────────────────────────────────────────────
    log('填入 BY店 去年同期…')
    ws_ly = wb['BY店 去年同期']
    fill_biz(ws_ly, 4, 'lymo', 'mtd')
    fill_units(ws_ly, 14, 'lymo', 'mtd')
    fill_acpp(ws_ly, 23, 'mtd');  fill_acpp(ws_ly, 32, 'lymo')

    # ── BY店 整年同期 ─────────────────────────────────────────────
    log('填入 BY店 整年同期…')
    ws_yoy = wb['BY店 整年同期']
    fill_biz(ws_yoy, 4, 'ytd_ly', 'ytd_cy')
    fill_units(ws_yoy, 14, 'ytd_ly', 'ytd_cy')
    fill_acpp(ws_yoy, 23, 'ytd_cy');  fill_acpp(ws_yoy, 32, 'ytd_ly')

    # ── BY店 本週其他細項 ─────────────────────────────────────────
    log('填入 BY店 本週其他細項…')
    ws_misc = wb['BY店 本週其他細項']
    for ri, code in enumerate(rows_stores):
        r = 3 + ri; m = MISC_WK[code]
        ws_misc.cell(r,  2).value = m['sa_mac'];      ws_misc.cell(r,  3).value = m['sa_iphone']
        ws_misc.cell(r,  4).value = m['sa_ipad'];     ws_misc.cell(r,  5).value = m['sa_watch']
        ws_misc.cell(r,  6).value = m['sa_airpods'];  ws_misc.cell(r,  7).value = m['sa_total']
        ws_misc.cell(r, 10).value = m['coupon_give']; ws_misc.cell(r, 11).value = m['coupon_redeem']
        ws_misc.cell(r, 12).value = rate(m['coupon_redeem'], m['coupon_give'])
        ws_misc.cell(r, 13).value = m['eco'];         ws_misc.cell(r, 14).value = m['cable']
        ws_misc.cell(r, 17).value = m['host_total']
    for ri, code in enumerate(rows_stores):
        r = 13 + ri; pw_m = MISC_PW[code]; wk_m = MISC_WK[code]
        ws_misc.cell(r, 2).value = pw_m['spk_with'];   ws_misc.cell(r, 3).value = wk_m['spk_with']
        ws_misc.cell(r, 4).value = pct(wk_m['spk_with'], pw_m['spk_with'])
        ws_misc.cell(r, 5).value = wk_m['spk_with'] - pw_m['spk_with']
        ws_misc.cell(r, 6).value = pw_m['spk_without']; ws_misc.cell(r, 7).value = wk_m['spk_without']
        ws_misc.cell(r, 8).value = pct(wk_m['spk_without'], pw_m['spk_without'])
        ws_misc.cell(r, 9).value = wk_m['spk_without'] - pw_m['spk_without']
    for ri, code in enumerate(rows_stores):
        r = 23 + ri; pw_m = MISC_PW[code]; wk_m = MISC_WK[code]
        ws_misc.cell(r,  2).value = pw_m['iphone_host'];     ws_misc.cell(r, 12).value = wk_m['iphone_host']
        ws_misc.cell(r,  3).value = pw_m['iphone_prot_qty']; ws_misc.cell(r, 13).value = wk_m['iphone_prot_qty']
        ws_misc.cell(r,  4).value = pw_m['iphone_case_qty']; ws_misc.cell(r, 14).value = wk_m['iphone_case_qty']
        ws_misc.cell(r,  5).value = pw_m['iphone_lens_qty']; ws_misc.cell(r, 15).value = wk_m['iphone_lens_qty']
        ws_misc.cell(r,  6).value = pw_m['iphone_prot_rev']; ws_misc.cell(r, 16).value = wk_m['iphone_prot_rev']
        ws_misc.cell(r,  7).value = pw_m['iphone_case_rev']; ws_misc.cell(r, 17).value = wk_m['iphone_case_rev']
        ws_misc.cell(r,  8).value = pw_m['iphone_lens_rev']; ws_misc.cell(r, 18).value = wk_m['iphone_lens_rev']
        ws_misc.cell(r,  9).value = rate(pw_m['iphone_prot_qty'], pw_m['iphone_host'])
        ws_misc.cell(r, 10).value = rate(pw_m['iphone_case_qty'], pw_m['iphone_host'])
        ws_misc.cell(r, 11).value = rate(pw_m['iphone_lens_qty'], pw_m['iphone_host'])
        ws_misc.cell(r, 19).value = rate(wk_m['iphone_prot_qty'], wk_m['iphone_host'])
        ws_misc.cell(r, 20).value = rate(wk_m['iphone_case_qty'], wk_m['iphone_host'])
        ws_misc.cell(r, 21).value = rate(wk_m['iphone_lens_qty'], wk_m['iphone_host'])
        ws_misc.cell(r, 22).value = wk_m['iphone_prot_qty']-pw_m['iphone_prot_qty']
        ws_misc.cell(r, 23).value = wk_m['iphone_case_qty']-pw_m['iphone_case_qty']
        ws_misc.cell(r, 24).value = wk_m['iphone_lens_qty']-pw_m['iphone_lens_qty']
        ws_misc.cell(r, 25).value = wk_m['iphone_prot_rev']-pw_m['iphone_prot_rev']
        ws_misc.cell(r, 26).value = wk_m['iphone_case_rev']-pw_m['iphone_case_rev']
        ws_misc.cell(r, 27).value = wk_m['iphone_lens_rev']-pw_m['iphone_lens_rev']
    for ri, code in enumerate(rows_stores):
        r = 33 + ri; pw_m = MISC_PW[code]; wk_m = MISC_WK[code]
        ws_misc.cell(r,  2).value = pw_m['ipad_host'];  ws_misc.cell(r, 13).value = wk_m['ipad_host']
        ws_misc.cell(r,  3).value = pw_m['ipad_pencil1']; ws_misc.cell(r, 14).value = wk_m['ipad_pencil1']
        ws_misc.cell(r,  4).value = rate(pw_m['ipad_pencil1'], pw_m['ipad_host'])
        ws_misc.cell(r,  5).value = pw_m['ipad_pencil3']; ws_misc.cell(r, 16).value = wk_m['ipad_pencil3']
        ws_misc.cell(r,  6).value = rate(pw_m['ipad_pencil3'], pw_m['ipad_host'])
        ws_misc.cell(r,  7).value = pw_m['ipad_prot_qty']; ws_misc.cell(r, 18).value = wk_m['ipad_prot_qty']
        ws_misc.cell(r,  8).value = rate(pw_m['ipad_prot_qty'], pw_m['ipad_host'])
        ws_misc.cell(r,  9).value = pw_m['ipad_case_qty']; ws_misc.cell(r, 20).value = wk_m['ipad_case_qty']
        ws_misc.cell(r, 10).value = rate(pw_m['ipad_case_qty'], pw_m['ipad_host'])
        ws_misc.cell(r, 11).value = pw_m['ipad_kb'];    ws_misc.cell(r, 22).value = wk_m['ipad_kb']
        ws_misc.cell(r, 12).value = rate(pw_m['ipad_kb'], pw_m['ipad_host'])
        ws_misc.cell(r, 15).value = rate(wk_m['ipad_pencil1'], wk_m['ipad_host'])
        ws_misc.cell(r, 17).value = rate(wk_m['ipad_pencil3'], wk_m['ipad_host'])
        ws_misc.cell(r, 19).value = rate(wk_m['ipad_prot_qty'], wk_m['ipad_host'])
        ws_misc.cell(r, 21).value = rate(wk_m['ipad_case_qty'], wk_m['ipad_host'])
        ws_misc.cell(r, 23).value = rate(wk_m['ipad_kb'], wk_m['ipad_host'])
    for ri, code in enumerate(rows_stores):
        r = 43 + ri; pw_m = MISC_PW[code]; wk_m = MISC_WK[code]
        ws_misc.cell(r, 2).value = pw_m['watch_host'];  ws_misc.cell(r, 7).value = wk_m['watch_host']
        ws_misc.cell(r, 3).value = pw_m['watch_prot'];  ws_misc.cell(r, 8).value = wk_m['watch_prot']
        ws_misc.cell(r, 4).value = rate(pw_m['watch_prot'], pw_m['watch_host'])
        ws_misc.cell(r, 5).value = pw_m['watch_band'];  ws_misc.cell(r,10).value = wk_m['watch_band']
        ws_misc.cell(r, 6).value = rate(pw_m['watch_band'], pw_m['watch_host'])
        ws_misc.cell(r, 9).value = rate(wk_m['watch_prot'], wk_m['watch_host'])
        ws_misc.cell(r,11).value = rate(wk_m['watch_band'], wk_m['watch_host'])

    # ── BY店 人員銷售 ─────────────────────────────────────────────
    log('填入 BY店 人員銷售…')
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.worksheet.cell_range import CellRange as _CellRange
    ws_staff = wb['BY店 人員銷售']
    STAFF_BLOCKS = [('004',1),('005',23),('024',44),('046',67),('054',93),('057',116)]

    def _insert_rows_fix(ws, idx, amount):
        merges = []
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= idx:
                merges.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
        heights = {}
        for r in range(idx, ws.max_row + 1):
            rd = ws.row_dimensions.get(r)
            if rd and rd.height:
                heights[r] = rd.height
        ws.insert_rows(idx, amount)
        for (min_r, max_r, min_c, max_c) in merges:
            try:
                ws.merged_cells.remove(
                    _CellRange(f'{_gcl(min_c)}{min_r}:{_gcl(max_c)}{max_r}'))
            except Exception:
                pass
        for (min_r, max_r, min_c, max_c) in merges:
            ws.merge_cells(start_row=min_r+amount, start_column=min_c,
                           end_row=max_r+amount, end_column=max_c)
        for r in heights:
            ws.row_dimensions[r].height = None
        for r, h in heights.items():
            ws.row_dimensions[r + amount].height = h

    _SKIP_LABELS = {
        '員工代碼', '員工名稱', '本週', '月累積', '小計',
        'iPhone台數', 'iPad台數', 'Watch台數', 'Mac台數', 'AirPods台數',
        'iPhone', 'iPad', 'Watch', 'Mac', 'AirPods', 'ACPP+', 'SAcare', '合計率',
    }

    def _find_blank_sec(ws, ds):
        slots, sub = [], None
        for r in range(ds, ds + 50):
            v1 = str(ws.cell(r, 1).value or '').strip()
            v2 = str(ws.cell(r, 2).value or '').strip()
            c3  = ws.cell(r, 3).value
            if v1 == '小計' or v2 == '小計':
                sub = r; break
            if v1 in _SKIP_LABELS or v2 in _SKIP_LABELS:
                continue
            if isinstance(c3, str) and c3.strip():   # 機型/欄位標頭行
                continue
            slots.append(r)
        return slots, sub

    def _copy_row_format(ws, src, dst):
        """將 src 列的格式複製到 dst 列（不複製值）。"""
        for c in range(1, ws.max_column + 1):
            sc_ = ws.cell(src, c); dc = ws.cell(dst, c)
            if sc_.has_style:
                dc.font        = _copy(sc_.font)
                dc.fill        = _copy(sc_.fill)
                dc.alignment   = _copy(sc_.alignment)
                dc.border      = _copy(sc_.border)
                dc.number_format = sc_.number_format
            dc.value = None
        src_h = ws.row_dimensions.get(src)
        if src_h and src_h.height:
            ws.row_dimensions[dst].height = src_h.height

    def _fill_person(ws, r, p):
        # col1=員工代碼  col2=員工名稱  col3~46=資料（新範本欄位佈局）
        ws.cell(r, 1).value = p.get('emp_id', '')
        ws.cell(r, 2).value = p.get('name', '')
        g = p.get
        # iPhone
        ws.cell(r, 3).value = g('iphone_host',0)
        ws.cell(r, 4).value = g('acpp_iphone',0)
        ws.cell(r, 5).value = g('sa_iphone',0)
        ws.cell(r, 6).value = rate(g('acpp_iphone',0)+g('sa_iphone',0), g('iphone_host',0))
        # iPad
        ws.cell(r, 7).value = g('ipad_host',0)
        ws.cell(r, 8).value = g('acpp_ipad',0)
        ws.cell(r, 9).value = g('sa_ipad',0)
        ws.cell(r,10).value = rate(g('acpp_ipad',0)+g('sa_ipad',0), g('ipad_host',0))
        # Watch
        ws.cell(r,11).value = g('watch_host',0)
        ws.cell(r,12).value = g('acpp_watch',0)
        ws.cell(r,13).value = g('sa_watch',0)
        ws.cell(r,14).value = rate(g('acpp_watch',0)+g('sa_watch',0), g('watch_host',0))
        # Mac
        ws.cell(r,15).value = g('mac_host',0)
        ws.cell(r,16).value = g('acpp_mac',0)
        ws.cell(r,17).value = g('sa_mac',0)
        ws.cell(r,18).value = rate(g('acpp_mac',0)+g('sa_mac',0), g('mac_host',0))
        # AirPods
        ws.cell(r,19).value = g('airpods_host',0)
        ws.cell(r,20).value = g('acpp_airpods',0)
        ws.cell(r,21).value = g('sa_airpods',0)
        ws.cell(r,22).value = rate(g('acpp_airpods',0)+g('sa_airpods',0), g('airpods_host',0))
        # iPhone 配件
        ws.cell(r,23).value = g('iphone_prot',0)
        ws.cell(r,24).value = g('iphone_case',0)
        ws.cell(r,25).value = g('iphone_lens',0)
        ws.cell(r,26).value = rate(g('iphone_prot',0), g('iphone_host',0))
        ws.cell(r,27).value = rate(g('iphone_case',0), g('iphone_host',0))
        ws.cell(r,28).value = rate(g('iphone_lens',0), g('iphone_host',0))
        # iPad 配件
        ws.cell(r,29).value = g('ipad_pencil1',0)
        ws.cell(r,30).value = g('ipad_pencil3',0)
        ws.cell(r,31).value = g('ipad_prot',0)
        ws.cell(r,32).value = g('ipad_case',0)
        ws.cell(r,33).value = g('ipad_kb',0)
        ws.cell(r,34).value = rate(g('ipad_pencil1',0), g('ipad_host',0))
        ws.cell(r,35).value = rate(g('ipad_pencil3',0), g('ipad_host',0))
        ws.cell(r,36).value = rate(g('ipad_prot',0),    g('ipad_host',0))
        ws.cell(r,37).value = rate(g('ipad_case',0),    g('ipad_host',0))
        ws.cell(r,38).value = rate(g('ipad_kb',0),      g('ipad_host',0))
        # Watch 配件
        ws.cell(r,39).value = g('watch_prot',0)
        ws.cell(r,40).value = g('watch_band',0)
        ws.cell(r,41).value = rate(g('watch_prot',0), g('watch_host',0))
        ws.cell(r,42).value = rate(g('watch_band',0), g('watch_host',0))
        # AirPods 配件
        ws.cell(r,43).value = g('airpods_acc',0)
        ws.cell(r,44).value = rate(g('airpods_acc',0), g('airpods_host',0))
        # 喇叭
        ws.cell(r,45).value = g('spk_with',0)
        ws.cell(r,46).value = g('spk_without',0)

    def _fill_sub(ws, r, m, misc):
        def g(k): return m.get(k) or 0
        ws.cell(r, 3).value = g('iphone_units')
        ws.cell(r, 4).value = g('acpp_iphone')
        ws.cell(r, 5).value = g('sa_iphone')
        ws.cell(r, 6).value = rate(g('acpp_iphone')+g('sa_iphone'), g('iphone_units'))
        ws.cell(r, 7).value = g('ipad_units')
        ws.cell(r, 8).value = g('acpp_ipad')
        ws.cell(r, 9).value = g('sa_ipad')
        ws.cell(r,10).value = rate(g('acpp_ipad')+g('sa_ipad'), g('ipad_units'))
        ws.cell(r,11).value = g('watch_units')
        ws.cell(r,12).value = g('acpp_watch')
        ws.cell(r,13).value = g('sa_watch')
        ws.cell(r,14).value = rate(g('acpp_watch')+g('sa_watch'), g('watch_units'))
        ws.cell(r,15).value = g('cpu_units')
        ws.cell(r,16).value = g('acpp_mac')
        ws.cell(r,17).value = g('sa_cpu')
        ws.cell(r,18).value = rate(g('acpp_mac')+g('sa_cpu'), g('cpu_units'))
        ws.cell(r,19).value = g('airpods_units')
        ws.cell(r,20).value = g('acpp_airpods')
        ws.cell(r,21).value = g('sa_airpods')
        ws.cell(r,22).value = rate(g('acpp_airpods')+g('sa_airpods'), g('airpods_units'))
        ws.cell(r,23).value = misc['iphone_prot_qty']
        ws.cell(r,24).value = misc['iphone_case_qty']
        ws.cell(r,25).value = misc['iphone_lens_qty']
        ws.cell(r,26).value = rate(misc['iphone_prot_qty'], m['iphone_units'])
        ws.cell(r,27).value = rate(misc['iphone_case_qty'], m['iphone_units'])
        ws.cell(r,28).value = rate(misc['iphone_lens_qty'], m['iphone_units'])
        ws.cell(r,29).value = misc['ipad_pencil1']
        ws.cell(r,30).value = misc['ipad_pencil3']
        ws.cell(r,31).value = misc['ipad_prot_qty']
        ws.cell(r,32).value = misc['ipad_case_qty']
        ws.cell(r,33).value = misc['ipad_kb']
        ws.cell(r,34).value = rate(misc['ipad_pencil1'],  m['ipad_units'])
        ws.cell(r,35).value = rate(misc['ipad_pencil3'],  m['ipad_units'])
        ws.cell(r,36).value = rate(misc['ipad_prot_qty'], m['ipad_units'])
        ws.cell(r,37).value = rate(misc['ipad_case_qty'], m['ipad_units'])
        ws.cell(r,38).value = rate(misc['ipad_kb'],       m['ipad_units'])
        ws.cell(r,39).value = misc['watch_prot']
        ws.cell(r,40).value = misc['watch_band']
        ws.cell(r,41).value = rate(misc['watch_prot'], m['watch_units'])
        ws.cell(r,42).value = rate(misc['watch_band'], m['watch_units'])
        ws.cell(r,45).value = misc.get('spk_with', 0)
        ws.cell(r,46).value = misc.get('spk_without', 0)

    def _fill_section(ws, ds, persons, m, misc):
        """填一個區段（本週或月累積）。回傳 (插入列數, 小計列號)。"""
        slots, sub = _find_blank_sec(ws, ds)
        if sub is None:
            return 0, ds
        inserted = 0
        if len(persons) > len(slots):
            n_extra = len(persons) - len(slots)
            ref = slots[-1] if slots else ds
            _insert_rows_fix(ws, sub, n_extra)
            for i in range(n_extra):
                _copy_row_format(ws, ref, sub + i)
                slots.append(sub + i)
            sub += n_extra
            inserted = n_extra
        for i, p in enumerate(persons):
            if i < len(slots):
                _fill_person(ws, slots[i], p)
        for r in slots[len(persons):]:
            for c in range(1, 47):
                ws.cell(r, c).value = None
        _fill_sub(ws, sub, m, misc)
        return inserted, sub

    row_offset = 0
    for sc, bs in STAFF_BLOCKS:
        actual_bs = bs + row_offset
        wk_persons  = eng.calc_person_metrics(df_cy, WK_START, wk_end,   sc, sa_prices)
        mtd_persons = eng.calc_person_metrics(df_cy, MTD_START, MTD_END, sc, sa_prices)
        log(f'  [{sc}] 本週 {len(wk_persons)} 人 / 月累積 {len(mtd_persons)} 人')
        ins_wk,  wk_sub  = _fill_section(ws_staff, actual_bs + 1, wk_persons,  M['wk'][sc],  MISC_WK[sc])
        ins_mtd, mtd_sub = _fill_section(ws_staff, wk_sub + 1,    mtd_persons, M['mtd'][sc], MISC_MTD[sc])
        row_offset += ins_wk + ins_mtd

    # ── 日期標題 ──────────────────────────────────────────────────
    log('更新日期標題…')

    def _d(d): return f'{d.month:02d}/{d.day:02d}'

    def _set_acc_dates(ws, off):
        ws.cell(3, off+3).value  = f'{_d(PW_START)}~{_d(PW_END)}'
        ws.cell(3, off+4).value  = f'{_d(WK_START)}~{_d(wk_end)}'
        ws.cell(3, off+7).value  = f'{_d(MTD_START)}~{_d(MTD_END)}'
        ws.cell(3, off+8).value  = f'{_d(PM_START)}~{_d(PM_END)}'
        ws.cell(3, off+11).value = f'{LYW_END.year}/{_d(LYW_START)}~{_d(LYW_END)}'
        ws.cell(3, off+14).value = f'{YTD_S_LY.year}/{_d(YTD_S_LY)}~{_d(YTD_E_LY)}'
        ws.cell(3, off+15).value = f'{YTD_S_CY.year}/{_d(YTD_S_CY)}~{_d(YTD_E_CY)}'

    for sname, label in [('配件-北一區 (匯總)','北一區'), ('配件-北一區','北一區'),
                          ('配件 - 士林門市','士林門市'), ('配件 - 微風門市','微風門市'),
                          ('配件 - 美麗華門市','美麗華門市'), ('配件 - 阿波羅門市','阿波羅門市'),
                          ('配件 - 大葉高島屋門市','大葉高島屋門市'), ('配件 - 羅東門市','羅東門市')]:
        if sname not in wb.sheetnames: continue
        ws_d = wb[sname]
        ws_d.cell(1,1).value = f'配件銷售分析  ·  {label}  ·  {WK_START.year}/{_d(WK_START)} ~ {_d(wk_end)}'
        _set_acc_dates(ws_d, 0)
        if sname == '配件-北一區 (匯總)': _set_acc_dates(ws_d, 17)

    ws_misc.cell(1,1).value = f'本週其他細項  ·  本週 {_d(WK_START)}~{_d(wk_end)}  |  對照 上週 {_d(PW_START)}~{_d(PW_END)}'

    # 配件區塊小標題（iPhone/iPad/Watch）：上週/本週日期
    _misc_pw = f'上週 {_d(PW_START)}~{_d(PW_END)}'
    _misc_wk = f'本週 {_d(WK_START)}~{_d(wk_end)}'
    for _r, _c_pw, _c_wk in [(21, 2, 12), (31, 2, 13), (41, 2, 7)]:
        ws_misc.cell(_r, _c_pw).value = _misc_pw
        ws_misc.cell(_r, _c_wk).value = _misc_wk

    def _rep_w(ws, row, la, lb):
        cnt = 0
        for c in range(1, ws.max_column+1):
            v = ws.cell(row, c).value
            if isinstance(v, str) and len(v) >= 2 and v[0]=='W' and v[1:].isdigit():
                cnt += 1
                ws.cell(row, c).value = la if cnt%2==1 else lb

    pw_lbl = f'{_d(PW_START)}~{_d(PW_END)}';  wk_lbl = f'{_d(WK_START)}~{_d(wk_end)}'
    _rep_w(ws_wk, 3, pw_lbl, wk_lbl);  _rep_w(ws_wk, 13, pw_lbl, wk_lbl)
    ws_wk.cell(22,1).value = wk_lbl;   ws_wk.cell(31,1).value = pw_lbl

    pm_s_lbl = f'{_d(PM_SAME_START)}~{_d(PM_SAME_END)}';  mtd_lbl = f'{_d(MTD_START)}~{_d(MTD_END)}'
    ws_mo.cell(1,1).value = f'月累積對照  ·  {MTD_START.year}/{mtd_lbl}  vs  {PM_SAME_START.year}/{pm_s_lbl}'
    for c in range(1, ws_mo.max_column+1):
        v = ws_mo.cell(3, c).value
        if not isinstance(v, str) or '\n' not in v: continue
        base = v.rsplit('\n', 1)[0]
        if c in {4,7,10,13,16,19}:   ws_mo.cell(3,c).value = base+'\n'+pm_s_lbl
        elif c in {5,8,11,14,17,20}: ws_mo.cell(3,c).value = base+'\n'+mtd_lbl
    pm_s_ymd=f'{PM_SAME_START.year}\n{pm_s_lbl}'; mtd_ymd=f'{MTD_START.year}\n{mtd_lbl}'
    for c in range(1, ws_mo.max_column+1):
        if ws_mo.cell(13,c).value is None: continue
        if c in {2,5,8,11,14}:  ws_mo.cell(13,c).value = pm_s_ymd
        elif c in {3,6,9,12,15}: ws_mo.cell(13,c).value = mtd_ymd
    ws_mo.cell(22,1).value = f'{MTD_START.year}\n{mtd_lbl}'
    ws_mo.cell(31,1).value = f'{PM_SAME_START.year}\n{pm_s_lbl}'

    ly_lbl = f'{_d(LYMO_START)}~{_d(LYMO_END)}'
    ws_ly.cell(1,1).value = f'去年同期對照  ·  本月 {MTD_START.year}/{mtd_lbl}  vs  去年同期 {LYMO_START.year}/{ly_lbl}'
    for c in range(1, ws_ly.max_column+1):
        v = ws_ly.cell(3,c).value
        if not isinstance(v, str) or '\n' not in v: continue
        parts = v.split('\n')
        if len(parts) >= 3:
            base = '\n'.join(parts[:-2])
            if c in {2,5,8,11,14,17}:   ws_ly.cell(3,c).value = f'{base}\n{LYMO_START.year}\n{ly_lbl}'
            elif c in {3,6,9,12,15,18}: ws_ly.cell(3,c).value = f'{base}\n{MTD_START.year}\n{mtd_lbl}'
    ly_ymd=f'{LYMO_START.year}\n{ly_lbl}'; cy_ymd=f'{MTD_START.year}\n{mtd_lbl}'
    for c in range(1, ws_ly.max_column+1):
        if ws_ly.cell(13,c).value is None: continue
        if c in {2,5,8,11,14}:  ws_ly.cell(13,c).value = ly_ymd
        elif c in {3,6,9,12,15}: ws_ly.cell(13,c).value = cy_ymd
    ws_ly.cell(22,1).value = f'{MTD_START.year}\n{mtd_lbl}'
    ws_ly.cell(31,1).value = f'{LYMO_START.year}\n{ly_lbl}'

    ytd_ly_lbl=f'{_d(YTD_S_LY)}~{_d(YTD_E_LY)}'; ytd_cy_lbl=f'{_d(YTD_S_CY)}~{_d(YTD_E_CY)}'
    for c in range(1, ws_yoy.max_column+1):
        v = ws_yoy.cell(3,c).value
        if not isinstance(v, str) or '\n' not in v: continue
        parts = v.split('\n')
        if len(parts) >= 3:
            base = '\n'.join(parts[:-2])
            if c in {2,5,8,11,14,17}:   ws_yoy.cell(3,c).value = f'{base}\n{YTD_S_LY.year}\n{ytd_ly_lbl}'
            elif c in {3,6,9,12,15,18}: ws_yoy.cell(3,c).value = f'{base}\n{YTD_S_CY.year}\n{ytd_cy_lbl}'
    ytd_ly_ymd=f'{YTD_S_LY.year}\n{ytd_ly_lbl}'; ytd_cy_ymd=f'{YTD_S_CY.year}\n{ytd_cy_lbl}'
    for c in range(1, ws_yoy.max_column+1):
        if ws_yoy.cell(13,c).value is None: continue
        if c in {2,5,8,11,14}:  ws_yoy.cell(13,c).value = ytd_ly_ymd
        elif c in {3,6,9,12,15}: ws_yoy.cell(13,c).value = ytd_cy_ymd
    ws_yoy.cell(22,1).value = f'{YTD_S_CY.year}\n{ytd_cy_lbl}'
    ws_yoy.cell(31,1).value = f'{YTD_S_LY.year}\n{ytd_ly_lbl}'

    log('儲存 Excel…')
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Job 管理 ─────────────────────────────────────────────────
def _run_job(job_id: str, wk_end_str: str, use_full_month: bool = False):
    def log(msg):
        ts = time.strftime('%H:%M:%S')
        with _LOCK:
            JOBS[job_id]['messages'].append(f'[{ts}] {msg}')

    with _LOCK:
        JOBS[job_id]['status'] = 'running'

    try:
        wk_end = date.fromisoformat(wk_end_str)
        result = _fill_workbook(wk_end, log, use_full_month=use_full_month)
        suffix = '_完整月' if use_full_month else ''
        filename = f'北一區週報_{wk_end}{suffix}.xlsx'
        with _LOCK:
            JOBS[job_id]['status']   = 'done'
            JOBS[job_id]['result']   = result
            JOBS[job_id]['filename'] = filename
        log(f'✓ 完成！({len(result):,} bytes)')
    except Exception:
        tb = traceback.format_exc()
        with _LOCK:
            JOBS[job_id]['status'] = 'error'
            JOBS[job_id]['error']  = tb
        log(f'✗ 錯誤:\n{tb}')


# ─── HTTP Handler ─────────────────────────────────────────────
class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_ROOT), **kwargs)

    def log_message(self, fmt, *args):
        pass  # suppress request logs

    def end_headers(self):
        self.send_header('Cache-Control', 'no-store')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def send_json(self, status, data):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_xlsx(self, filename, body):
        quoted = urllib.parse.quote(filename)
        self.send_response(200)
        self.send_header('Content-Type',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition',
            f"attachment; filename*=UTF-8''{quoted}")
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_body(self):
        n = int(self.headers.get('Content-Length', 0))
        return json.loads(self.rfile.read(n).decode('utf-8') if n else '{}')

    def do_GET(self):
        p = urllib.parse.urlparse(self.path)
        qs = urllib.parse.parse_qs(p.query)

        if p.path == '/api/default-date':
            self.send_json(200, {'date': _last_saturday().isoformat()})
            return

        if p.path == '/api/traffic-status':
            stores = _load_traffic()
            all_dates = sorted(d for days in stores.values() for d in days)
            try:
                with open(TRAFFIC_FILE, encoding='utf-8') as f:
                    updated = json.load(f).get('updated', '')
            except Exception:
                updated = ''
            self.send_json(200, {
                'storeCount': len(stores),
                'codes':      sorted(stores.keys()),
                'latest':     all_dates[-1] if all_dates else '',
                'earliest':   all_dates[0] if all_dates else '',
                'updated':    updated,
            })
            return

        if p.path == '/api/status':
            job_id = qs.get('jobId', [''])[0]
            with _LOCK:
                job = JOBS.get(job_id)
            if not job:
                self.send_json(404, {'error': '找不到工作'})
                return
            resp = {
                'status':   job['status'],
                'messages': list(job['messages']),
                'filename': job.get('filename', ''),
            }
            if job['status'] == 'error':
                resp['error'] = job.get('error', '未知錯誤')
            self.send_json(200, resp)
            return

        if p.path == '/api/download':
            job_id = qs.get('jobId', [''])[0]
            with _LOCK:
                job = JOBS.get(job_id)
            if not job or job['status'] != 'done':
                self.send_json(400, {'error': '檔案尚未就緒'})
                return
            self.send_xlsx(job['filename'], job['result'])
            return

        super().do_GET()

    def do_POST(self):
        p = urllib.parse.urlparse(self.path)
        if p.path == '/api/generate':
            try:
                payload       = self.read_body()
                wk_end_s      = str(payload.get('week_end', payload.get('wkEnd', ''))).strip()
                use_full_month = bool(payload.get('useFullMonth', False))
                date.fromisoformat(wk_end_s)  # validate
            except Exception as e:
                self.send_json(400, {'error': f'日期格式錯誤: {e}'})
                return
            job_id = str(uuid.uuid4())
            with _LOCK:
                JOBS[job_id] = {'status': 'pending', 'messages': [], 'result': None}
            threading.Thread(target=_run_job, args=(job_id, wk_end_s, use_full_month),
                             daemon=True).start()
            self.send_json(200, {'jobId': job_id})
            return

        if p.path == '/api/traffic':
            # 插件背景推送人流：{ "stores": { "004": {"2026-05-24": 123, ...}, ... } }
            try:
                payload = self.read_body()
                stores  = payload.get('stores', {})
                if not isinstance(stores, dict):
                    raise ValueError('stores 格式錯誤')
                # 合併（同店同日覆寫，保留其他日期）
                cur = _load_traffic()
                for code, days in stores.items():
                    cur.setdefault(str(code), {}).update(
                        {str(d): int(v) for d, v in days.items()})
                _save_traffic(cur)
                n_days = sum(len(v) for v in stores.values())
            except Exception as e:
                self.send_json(400, {'error': f'人流資料錯誤: {e}'})
                return
            self.send_json(200, {'ok': True, 'stores': len(stores), 'days': n_days})
            return

        self.send_json(404, {'error': 'Not found'})


def main():
    port = int(os.environ.get('PORT', '8782'))
    server = ThreadingHTTPServer(('127.0.0.1', port), Handler)
    print(f'北一區週報產生器：http://127.0.0.1:{port}', flush=True)
    print('按 Ctrl+C 停止', flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('已停止。')


if __name__ == '__main__':
    main()
